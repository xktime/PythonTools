import json
import redis 
import openpyxl as op

r = redis.Redis(host='r-bp1omf2u1x723gngds.redis.rds.aliyuncs.com',port=6379 password='Qcplay!@#%')
wb = op.Workbook()
sheet = wb.create_sheet("导出数据", 0)

length = 0
for k, v in r.hgetall('ST_COMBAT').items():
    v = json.loads(v.decode())
    length += 1
    sheet.cell(length, 1).value = k
    sheet.cell(length, 2).value = v["attack_id"]
    sheet.cell(length, 3).value = v["denfend_id"]
    sheet.cell(length, 4).value = v["combat_expend_ms"]
    sheet.cell(length, 5).value = v["cost_time"]
    sheet.cell(length, 6).value = v["cur_times"]
    sheet.cell(length, 7).value = str(v)

wb.save("./data.xlsx")