import openpyxl as op
import time
from elasticsearch import Elasticsearch
import xml.dom.minidom as xml

ROLE_ID_COLUMN = 1
ROLE_NAME_COLUMN = 2
LEVEL_COLUMN = 3
GUIDE_ID_COLUMN = 4
CHAPTER_COLUMN = 5
LOGIN_TIME_COLUMN = 6


def get_es_data(index, count):
    dom = xml.parse('config.xml')
    root = dom.documentElement
    address = root.getElementsByTagName('address')[0].firstChild.data
    username = root.getElementsByTagName('username')[0].firstChild.data
    password = root.getElementsByTagName('password')[0].firstChild.data
    es = Elasticsearch(['http://' + address], http_auth=(username, password))
    fromIndex = 0
    size = 1000
    results = []
    while fromIndex < count:
        body = {
            "query": {
                "match_all": {
                },
            },
            "from": fromIndex,
            "size": size,
            "track_total_hits": "true"
        }
        fromIndex += size
        datas = es.search(index=index, body=body)
        results.extend(datas['hits']['hits'])
    return results


def export_to_excel(index, count):
    wb = op.Workbook()
    sheet = wb.create_sheet("导出数据", 0)
    datas = get_es_data(index, count)
    indexMapping = {}
    for i in range(1, len(datas)):
        data = datas.__getitem__(i)
        if data is None:
            continue
        info = data["_source"]
        if i == 1:
            keys = list(info.keys())
            for j in range(1, len(keys)):
                key = keys.__getitem__(j)
                sheet.cell(1, j).value = key
                indexMapping[key] = j
        for key, value in info.items():
            if not indexMapping.__contains__(key):
                continue
            sheet.cell(i + 1, indexMapping.__getitem__(key)).value = str(value)

    wb.save("./result.xlsx")


print("请输入需要查询的索引")
index = input()
print("请输入需要查询多少条数据")
count = input()

try:
    count = int(count)
except:
    print("请输入数字")

try:
    export_to_excel(index, count)
except Exception as e:
    print(e)
