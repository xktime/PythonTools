import datetime
import os
import re

import openpyxl as op

'''
----------------------------------------实体类------------------------------
'''


class TargetEntity:
    def __init__(self):
        self.map = {}
        self.companyId = 0
        self.time = 0
        self.companyName = ""
        self.index = 0


TIME_COLUMN = 1
COMPANY_NAME_COLUMN = 2
INDEX_COLUMN = 3
COMPANY_ID_COLUMN = 4
'''
----------------------------------------method------------------------------
'''


def find_month_datas(entitys, sheet, year_column, year_row):
    month = 0
    year = sheet.cell(year_row, year_column).value.replace('\n', '')
    if year not in entitys:
        monthMap = {}
        entitys[year] = monthMap
    monthMap = entitys[year]
    for dataColumn in range(year_column, year_column + 12):
        month += 1
        if month not in monthMap:
            dataMap = {}  ##map<companyId, List<TargetEntity>>
            monthMap[month] = dataMap
        dataMap = monthMap[month]
        for row in range(year_row + 1, sheet.max_row):
            index = sheet.cell(row, 1).value
            if index is None:
                continue
            companyId = sheet.cell(row, 2).value
            if companyId not in dataMap:
                entity = TargetEntity()
                entity.companyName = sheet.cell(row, 4).value
                entity.companyId = companyId
                entity.index = index
                entity.time = str(year) + str(month) + "月"
                dataMap[companyId] = entity
            entity = dataMap[companyId]
            projectName = sheet.cell(row, 5).value
            if projectName in entity.map and entity.map[projectName] > 0:
                continue
            dataValue = sheet.cell(row, dataColumn).value
            entity.map[projectName] = {True: 0, False: dataValue}[dataValue is None]


##return map<year, map<month, map<companyId, TargetEntity>>
def get_entitys(source_path):
    sourceWb = op.load_workbook(source_path)
    sheetNames = sourceWb.sheetnames
    entitys = {}
    for sheetName in sheetNames:
        sheet = sourceWb[sheetName]
        if re.match(r'.*税率*', str(sheetName)):
            continue
        max_row = sheet.max_row
        max_column = sheet.max_column
        for column in range(1, max_column):
            for row in range(1, max_row):
                cell = sheet.cell(row, column)
                year = cell.value
                if year is None:
                    continue
                if re.match(r'(\d){4}年', str(year)) is None:
                    continue
                find_month_datas(entitys, sheet, column, row)
    return entitys


def create_and_write_excel_file(entitys, file_path):
    if os.path.exists(file_path):
        print("该文件已存在，是否覆盖当前文件，输入yes or no")
        isDelete = input()
        if isDelete == "yes":
            os.remove(file_path)
        else:
            nowTime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            file_path = './' + nowTime + '.xlsx'
    wb = op.Workbook()
    for year, monthMap in entitys.items():
        for month, entityMap in monthMap.items():
            sheet = wb.create_sheet(str(year) + str(month) + "月")
            entityList = list(entityMap.values())
            for row in range(1, len(entityList) + 1):
                if row == 1:
                    sheet.cell(row, TIME_COLUMN).value = '时间'
                    sheet.cell(row, COMPANY_NAME_COLUMN).value = '项目名称'
                    sheet.cell(row, INDEX_COLUMN).value = '行号'
                    sheet.cell(row, COMPANY_ID_COLUMN).value = '发电客户编号'
                    columnNum = COMPANY_ID_COLUMN + 1
                    for columnName in entityList[0].map.keys():
                        sheet.cell(row, columnNum).value = str(columnName)
                        columnNum += 1
                else:
                    entity = entityList[row - 2]
                    sheet.cell(row, TIME_COLUMN).value = entity.time
                    sheet.cell(row, COMPANY_NAME_COLUMN).value = entity.companyName
                    sheet.cell(row, INDEX_COLUMN).value = entity.index
                    sheet.cell(row, COMPANY_ID_COLUMN).value = entity.companyId
                    for column in range(COMPANY_ID_COLUMN + 1, sheet.max_column + 1):
                        dataName = sheet.cell(1, column).value
                        dataMap = entity.map
                        if dataName in dataMap:
                            sheet.cell(row, column).value = dataMap[dataName]
    wb.save(file_path)


'''
----------------------------------------main------------------------------
'''
print("请输入需要转换的文件名")
source_file_name = './' + input() + '.xlsx'
print("请输入需要生成的文件名")
target_file_path = './' + input() + '.xlsx'
create_and_write_excel_file(get_entitys(source_file_name), target_file_path)
